import pandas as pd
import openpyxl
import glob
from fpdf import FPDF
from pathlib import Path

sheets = glob.glob('sheets/*.xlsx')
print(sheets)

for sheet in sheets:

    filename = Path(sheet).stem
    invoice_nr, date = filename.split("-")

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=False, margin=0)
    pdf.add_page()

    pdf.set_font(family="Courier", style="B", size=16)
    pdf.cell(w=50, h=35, txt=f"invoice_nr.{invoice_nr}", align="L",
             ln=1)

    pdf.set_font(family="Courier", style="B", size=16)
    pdf.cell(w=50, h=8, txt=f"Date: {date}", align="L", ln=1)

    df = pd.read_excel(sheet, sheet_name='Sheet 1')

# Add Header
    cols = df.columns
    cols = [item.replace("_", " ").title() for item in cols]

    pdf.set_font(family="Times", style="B", size=10)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=28, h=8, txt=cols[0], border=1)
    pdf.cell(w=60, h=8, txt=cols[1], border=1)
    pdf.cell(w=38, h=8, txt=cols[2],
             border=1)
    pdf.cell(w=32, h=8, txt=cols[3], border=1)
    pdf.cell(w=32, h=8, txt=cols[4], border=1,
             ln=1)
# Add Rows
    for index, row in df.iterrows():
        pdf.set_font(family="Courier", style="B", size=10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(w=28, h=8, txt=str(row["product_id"]), border=1)
        pdf.cell(w=60, h=8, txt=str(row["product_name"]), border=1)
        pdf.cell(w=38, h=8, txt=str(row["amount_purchased"]),
                 border=1)
        pdf.cell(w=32, h=8, txt=f'$ {str(row["price_per_unit"])}.00', border=1)
        pdf.cell(w=32, h=8, txt=f'$ {str(row["total_price"])}.00', border=1,
                 ln=1)

    total_sum = df['total_price'].sum()
    pdf.set_font(family="Courier", style="B", size=10)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=28, h=8, txt="", border=1)
    pdf.cell(w=60, h=8, txt="", border=1)
    pdf.cell(w=38, h=8, txt="",
             border=1)
    pdf.cell(w=32, h=8, txt="",
             border=1)
    pdf.cell(w=32, h=8, txt=f'$ {str(total_sum)}.00',
             border=1,
             ln=1)

    pdf.set_font(family="Courier", style="B", size=10)
    pdf.cell(w=32, h=8, txt=f' Total price: ${str(total_sum)}.00', ln=1)

    pdf.set_font(family="Courier", style="B", size=30)
    pdf.cell(w=32, h=8, txt=f'chainai.online', ln=1)
    pdf.image('logo.png', w=70)
    pdf.output(f"pdf/{filename}.pdf")
